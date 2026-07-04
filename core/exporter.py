import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FIELDS = ["provider_name", "town_or_location", "address", "website", "email", "phone", "source_url", "email_source", "notes"]


def export(records: list[dict], output_dir: str | Path, basename: str = "citb_providers") -> tuple[Path, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path, xlsx_path = output_dir / f"{basename}.csv", output_dir / f"{basename}.xlsx"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
    wb = Workbook()
    ws = wb.active
    ws.title = "CITB Providers"
    ws.append(FIELDS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for record in records:
        ws.append([record.get(field, "") for field in FIELDS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, field in enumerate(FIELDS, 1):
        width = min(60, max(len(field) + 2, *(len(str(ws.cell(row, index).value or "")) + 2 for row in range(2, ws.max_row + 1))))
        ws.column_dimensions[get_column_letter(index)].width = width
    wb.save(xlsx_path)
    return csv_path, xlsx_path
